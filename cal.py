"""Risky value Calculator

This is a module to calculate the risky value

"""
import openpyxl

wb = openpyxl.load_workbook("./value.xlsx")
sheet_list = wb.sheetnames

T_SHEET = wb[sheet_list[1]]
V_SHEET = wb[sheet_list[2]]
T_MAX_ROW = T_SHEET.max_row
T_MAX_COLUNM = T_SHEET.max_column
V_MAX_ROW = V_SHEET.max_row
V_MAX_COLUNM = V_SHEET.max_column

t_list = []
v_list = []

for i in range(T_MAX_ROW):
    t_name_value = T_SHEET.cell(i + 1, 3).value
    t_threat_value = T_SHEET.cell(i + 1, T_MAX_COLUNM).value
    if i+1 == 1:
        pass
    else:
        t_tuple = (t_name_value, t_threat_value)
        t_list.append(t_tuple)

for i in range(V_MAX_ROW):
    v_name_value = V_SHEET.cell(i + 1, 2).value
    v_threat_value = V_SHEET.cell(i + 1, V_MAX_COLUNM).value
    if i+1 == 1:
        pass
    elif v_name_value != None:
        temp = v_name_value
        v_tuple = (v_name_value, v_threat_value)
        v_list.append(v_tuple)
    elif v_name_value == None:
        v_name_value = temp
        v_tuple = (v_name_value, v_threat_value)
        v_list.append(v_tuple)
    else:
        print("Something is wrong.")

# print(t_list)
# print(v_list)
