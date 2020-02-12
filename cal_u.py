"""Risky value Calculator

This is a module to calculate the risky value

"""

import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font


def get_arguments():
    """Parsing the arguments"""
    parser = argparse.ArgumentParser()
    parser.add_argument("-s",
                        "--source",
                        dest="source",
                        help="Set source file path for the script to parse.")
    parser.add_argument("-o",
                        "--output",
                        dest="output",
                        help="Set output file path.")
    options = parser.parse_args()
    return options


def sheet_format(sheet):
    """Add header to the output sheet & set bold font, alignment, cell size, background color"""
    header = ["资产名称", "资产赋值", "威胁等级", "脆弱性赋值", "已有安全措施", "风险值"]
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16
    sheet.row_dimensions[1].height = 20
    sheet.insert_rows(1)
    font = Font(bold=True)
    fill = PatternFill(fill_type="solid",
                       start_color="D9D9D9",
                       end_color="D9D9D9")
    alignment = Alignment(horizontal="center", vertical="center")
    for i in range(len(header)):
        sheet.cell(1, i + 1).value = header[i]
        sheet.cell(1, i + 1).font = font
        sheet.cell(1, i + 1).fill = fill
        sheet.cell(1, i + 1).alignment = alignment


def calculate(sheet, index):
    """calculate the risk value"""
    if index == 1:
        pass
    else:
        formula = f"=B{index}*SQRT(C{index}*D{index})*(5-E{index})"
        sheet["F" + str(index-1)] = formula


# get args
options = get_arguments()
# load a workbook -> .xlsx file
wb = openpyxl.load_workbook(options.source, data_only=True)
sheet_list = wb.sheetnames
# extract threat & vulnerability sheet
A_SHEET = wb[sheet_list[0]]
T_SHEET = wb[sheet_list[1]]
V_SHEET = wb[sheet_list[2]]
M_SHEET = wb[sheet_list[3]]
# set max row and column
A_MAX_ROW = A_SHEET.max_row  # 46
A_MAX_COLUMN = A_SHEET.max_column  # 7
T_MAX_ROW = T_SHEET.max_row
T_MAX_COLUMN = T_SHEET.max_column
V_MAX_ROW = V_SHEET.max_row
V_MAX_COLUMN = V_SHEET.max_column
M_MAX_ROW = M_SHEET.max_row
M_MAX_COLUMN = M_SHEET.max_column

t_list = []
v_list = []
asset_list = []
asset_flag = 0  # to save the start number of a_list
t_dic = {}
v_dic = {}
t_flag = ""
v_flag = ""
i_tuple = ()
a_list = []

# new workbook for output
output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = "Integration"

# Parse Asset Value
for row in A_SHEET.iter_rows(2, A_MAX_ROW, 1, A_MAX_COLUMN, values_only=True):
    if row[0] == None:
        del row
    else:
        asset_list.append((row[1], row[-1]))

# Parse Threat Value
for i in range(T_MAX_ROW):
    t_name_value = T_SHEET.cell(i + 1, 3).value  # name value
    t_threat_value = T_SHEET.cell(i + 1, T_MAX_COLUMN).value  # threat value
    if i + 1 == 1:  # skip the header
        pass
    elif t_flag == "":
        t_flag = t_name_value
        t_list.append(t_threat_value)
        t_dic[t_name_value] = t_list
    elif t_name_value == t_flag:
        t_list.append(t_threat_value)
        t_dic[t_name_value] = t_list
    else:
        t_flag = t_name_value
        t_list = []
        t_list.append(t_threat_value)
        t_dic[t_name_value] = t_list

# Parse Vulnerability Value
for i in range(V_MAX_ROW):
    v_name_value = V_SHEET.cell(i + 1, 2).value
    v_threat_value = V_SHEET.cell(i + 1, 5).value
    m_mitigate = M_SHEET.cell(i + 1, M_MAX_COLUMN).value
    if i + 1 == 1:
        pass
    elif v_name_value == None:
        v_name_value = v_flag
        v_list.append((v_threat_value, m_mitigate))
        v_dic[v_name_value] = v_list
    else:
        v_flag = v_name_value
        v_list = []
        v_list.append((v_threat_value, m_mitigate))
        v_dic[v_name_value] = v_list

# integrate threat & vulnerability
for name in t_dic:
    for t in t_dic[name]:
        for v in v_dic[name]:
            i_tuple = (name, t, v[0], v[1])
            a_list.append(i_tuple)

for t in asset_list:
    for i in range(asset_flag, len(a_list)):
        if t[0] == a_list[i][0]:
            output_sheet.cell(row=i + 1, column=1, value=a_list[i][0])
            output_sheet.cell(row=i + 1, column=2, value=t[1])
            output_sheet.cell(row=i + 1, column=3, value=a_list[i][1])
            output_sheet.cell(row=i + 1, column=4, value=a_list[i][2])
            output_sheet.cell(row=i + 1, column=5, value=a_list[i][3])
        else:
            asset_flag = i
            break

for i in range(len(a_list) + 1):
    calculate(output_sheet, i + 1)

sheet_format(output_sheet)
output_wb.save(options.output)
