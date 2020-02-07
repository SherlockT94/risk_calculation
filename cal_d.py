"""Risky value Calculator

This is a module to calculate the risky value

"""

import openpyxl
from openpyxl import Workbook

# load a workbook -> .xlsx file
wb = openpyxl.load_workbook("./value.xlsx", data_only=True)
sheet_list = wb.sheetnames
# extract threat & vulnerability sheet
A_SHEET = wb[sheet_list[0]]
T_SHEET = wb[sheet_list[1]]
V_SHEET = wb[sheet_list[2]]
# set max row and column
A_MAX_ROW = A_SHEET.max_row # 46
A_MAX_COLUMN = A_SHEET.max_column # 7
T_MAX_ROW = T_SHEET.max_row
T_MAX_COLUMN = T_SHEET.max_column
V_MAX_ROW = V_SHEET.max_row
V_MAX_COLUMN = V_SHEET.max_column

t_list = []
v_list = []
asset_list = []
t_dic = {}
v_dic = {}
t_flag = ""
v_flag = ""
i_tuple = ()
a_list = []

# new workbook for output
output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = "Integration"

for row in A_SHEET.iter_rows(2, A_MAX_ROW, 1, A_MAX_COLUMN, values_only=True):
    if row[0] == None:
        del row
    else:
        asset_list.append((row[1], row[-1]))
print(asset_list)
        
for i in range(T_MAX_ROW):
    t_name_value = T_SHEET.cell(i + 1, 3).value
    t_threat_value = T_SHEET.cell(i + 1, T_MAX_COLUMN).value
    if i + 1 == 1:
        pass
    elif t_flag == "":
        t_flag = t_name_value
        t_list.append(t_threat_value)
        t_dic[t_name_value] = t_list
    elif t_name_value == t_flag:
        t_list.append(t_threat_value)
        t_dic[t_name_value] = t_list
    else:
        t_flag = t_name_value
        t_list = []
        t_list.append(t_threat_value)
        t_dic[t_name_value] = t_list

for i in range(V_MAX_ROW):
    v_name_value = V_SHEET.cell(i + 1, 2).value
    v_threat_value = V_SHEET.cell(i + 1, V_MAX_COLUMN).value
    if i + 1 == 1:
        pass
    elif v_name_value == None:
        v_name_value = v_flag
        v_list.append(v_threat_value)
        v_dic[v_name_value] = v_list
    else:
        v_flag = v_name_value
        v_list = []
        v_list.append(v_threat_value)
        v_dic[v_name_value] = v_list

# integrate threat & vulnerability
for name in t_dic:
    for t in t_dic[name]:
        for v in v_dic[name]:
            i_tuple = (name, t, v)
            a_list.append(i_tuple)

for i in range(len(a_list)):
    output_sheet.cell(row=i+1, column=1, value=a_list[i][0])
    output_sheet.cell(row=i+1, column=2, value=a_list[i][1])
    output_sheet.cell(row=i+1, column=3, value=a_list[i][2])

# output_wb.save('output.xlsx')
